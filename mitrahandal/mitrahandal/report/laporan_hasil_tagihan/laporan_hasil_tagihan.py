# Copyright (c) 2026, Mitrahandal and contributors
# For license information, please see license.txt

import frappe
from frappe import _
from collections import defaultdict
from frappe.utils import getdate, format_date, fmt_money, nowdate, flt, now_datetime
import os
from datetime import datetime
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Setup logging to mitrahandal/logs folder
log_dir = os.path.join(os.path.dirname(__file__), "..", "..", "..", "logs")
os.makedirs(log_dir, exist_ok=True)
log_file = os.path.join(log_dir, "laporan_hasil_tagihan-" + now_datetime().strftime("%Y-%m-%d") + ".log")


def log_debug(message):
    timestamp = now_datetime().strftime("%Y-%m-%d %H:%M:%S")
    with open(log_file, "a") as f:
        f.write(f"[{timestamp}] {message}\n")


def execute(filters=None):
    columns = get_columns()
    data = get_data(filters)

    return columns, data


def validate_filters(filters):
    pass


@frappe.whitelist()
def get_sales_persons_by_branch(branch):
    """
    Get list of Sales Persons from unpaid Sales Invoices for a specific Branch.
    Returns a list of Sales Person names.
    """
    if not branch:
        return []
    
    # Get all unpaid Sales Invoices for customers in the specified branch
    # First, get customers that belong to this branch
    customers = frappe.get_all(
        "Customer",
        filters={"custom_branch": branch},
        fields=["name"]
    )
    
    if not customers:
        return []
    
    customer_list = [c.name for c in customers]
    
    # Get unpaid Sales Invoices for these customers
    sales_invoices = frappe.get_all(
        "Sales Invoice",
        filters={
            "customer": ["in", customer_list],
            "status": "Unpaid",
            "docstatus": 1
        },
        fields=["name"]
    )
    
    if not sales_invoices:
        return []
    
    # Get Sales Persons from Sales Invoice (assuming there's a sales_person field)
    # Check if Sales Team child table exists
    sales_persons = set()
    
    for inv in sales_invoices:
        # Try to get from Sales Team child table
        sales_team = frappe.get_all(
            "Sales Team",
            filters={"parenttype": "Sales Invoice", "parent": inv.name},
            fields=["sales_person"]
        )
        
        for st in sales_team:
            if st.sales_person:
                sales_persons.add(st.sales_person)
    
    # If no sales persons found from Sales Team, try to get from sales_person field directly
    if not sales_persons:
        for inv in sales_invoices:
            inv_doc = frappe.get_doc("Sales Invoice", inv.name)
            if hasattr(inv_doc, 'sales_person') and inv_doc.sales_person:
                sales_persons.add(inv_doc.sales_person)
    
    return sorted(list(sales_persons))


def get_columns():
    return [
        {
            "label": "Customer",
            "fieldname": "customer",
            "fieldtype": "Link",
            "options": "Customer",
            "width": 150
        },
        {
            "label": "Customer Name",
            "fieldname": "customer_name",
            "fieldtype": "Data",
            "width": 200
        },
        {
            "label": "Doc. No",
            "fieldname": "custom_doc_no",
            "fieldtype": "Data",
            "width": 120
        },
        {
            "label": "Inv. Date",
            "fieldname": "posting_date",
            "fieldtype": "Date",
            "width": 120
        },
        {
            "label": "Exp. Date",
            "fieldname": "due_date",
            "fieldtype": "Date",
            "width": 120
        },
        {
            "label": "Amount Invoice",
            "fieldname": "paid_amount",
            "fieldtype": "Currency",
            "width": 140
        },
        {
            "label": "Balance Due",
            "fieldname": "outstanding_amount",
            "fieldtype": "Currency",
            "width": 140
        },
        {
            "label": "Mode Of Payment",
            "fieldname": "mode_of_payment",
            "fieldtype": "Data",
            "width": 150
        },
        {
            "label": "Payment Amount",
            "fieldname": "payment_amount",
            "fieldtype": "Currency",
            "width": 140
        },
        {
            "label": "Tgl Bayar",
            "fieldname": "payment_date",
            "fieldtype": "Date",
            "width": 120
        },
    ]


def get_data(filters):
    log_debug(f"Filters received: {filters} Laporan Hasil Tagihan Debug")
    validate_filters(filters)

    # Build filters for frappe.get_all
    si_filters = {"docstatus": 1, "status": "Unpaid"}

    # Filter Company
    if filters.get("company"):
        si_filters["company"] = filters.get("company")

    log_debug(f"Base SI filters: {si_filters}")

    # Get Sales Invoice list
    invoice_list = frappe.get_all(
        "Sales Invoice",
        filters=si_filters,
        fields=[
            "name",
            "customer",
            "customer_name",
            "custom_doc_no",
            "posting_date",
            "due_date",
            "grand_total",
            "outstanding_amount",
        ],
        order_by="posting_date DESC, name DESC",
    )
    log_debug(f"Initial invoice list: {len(invoice_list)} invoices")

    # Filter by Branch: Sales Person's custom_branch
    # Get Sales Persons that belong to the selected branch
    if filters.get("branch"):
        branch_sales_persons = frappe.get_all(
            "Sales Person",
            filters={"custom_branch": filters.get("branch")},
            fields=["name"]
        )
        branch_sp_names = [sp.name for sp in branch_sales_persons]
        log_debug(f"Branch '{filters.get('branch')}' Sales Persons: {branch_sp_names}")

        if branch_sp_names:
            filtered_invoices = []
            for inv in invoice_list:
                sales_team = frappe.get_all(
                    "Sales Team",
                    filters={"parenttype": "Sales Invoice", "parent": inv.name},
                    fields=["sales_person"]
                )
                matched = any(st.sales_person in branch_sp_names for st in sales_team if st.sales_person)
                if matched:
                    filtered_invoices.append(inv)
            invoice_list = filtered_invoices
            log_debug(f"After branch filter: {len(invoice_list)} invoices")

    # Filter by Sales Person (ARRAY): fieldtype Data, could be comma-separated or JSON string
    sales_persons = filters.get("sales_persons")
    if sales_persons:
        if isinstance(sales_persons, str):
            # Try JSON first (could be '["SP1", "SP2"]')
            try:
                sales_persons = json.loads(sales_persons)
            except:
                # Fallback: comma-separated
                sales_persons = [sp.strip() for sp in sales_persons.split(",") if sp.strip()]
        if not isinstance(sales_persons, list):
            sales_persons = [sales_persons]

        sales_persons = [sp.strip() for sp in sales_persons if sp.strip()]

        if sales_persons:
            log_debug(f"Sales persons filter: {sales_persons}")

            filtered_invoices = []
            for inv in invoice_list:
                # Check Sales Team child table - match ANY sales person in the array
                sales_team = frappe.get_all(
                    "Sales Team",
                    filters={"parenttype": "Sales Invoice", "parent": inv.name},
                    fields=["sales_person"]
                )

                # Check if any sales team member is in the selected sales_persons list
                matched = any(st.sales_person in sales_persons for st in sales_team if st.sales_person)

                if matched:
                    filtered_invoices.append(inv)

            invoice_list = filtered_invoices
            log_debug(f"After sales persons filter: {len(invoice_list)} invoices")

    # Filter by routing: match customer's custom_routing with sdate
    if filters.get("sdate"):
        sdate = getdate(filters.get("sdate"))

        def get_week_mod(tanggal):
            weeknum = tanggal.isocalendar()[1]  # ISO week number (1-53)
            mod_val = weeknum % 4
            return 4 if mod_val == 0 else mod_val

        week_cycle = get_week_mod(sdate)
        week_cycle_str = str(week_cycle)
        day_digit = str(sdate.isoweekday())  # 1=Senin, 7=Minggu

        log_debug(f"Routing filter: day_digit={day_digit}, week_cycle={week_cycle}")

        filtered_invoices = []
        for inv in invoice_list:
            custom_routing = frappe.db.get_value("Customer", inv.customer, "custom_routing")
            if not custom_routing:
                continue

            routing_suffix = custom_routing[-3:]  # last 3 digits
            routing_day = routing_suffix[0]      # 1st digit = day
            routing_weeks = routing_suffix[1:]   # 2nd & 3rd digit = weeks

            if routing_day == day_digit and week_cycle_str in routing_weeks:
                filtered_invoices.append(inv)

        invoice_list = filtered_invoices
        log_debug(f"After routing filter: {len(invoice_list)} invoices")

    data = []
    for inv in invoice_list:
        # Get Payment Entry references that link to this Sales Invoice
        payment_references = frappe.get_all(
            "Payment Entry Reference",
            filters={
                "reference_doctype": "Sales Invoice",
                "reference_name": inv.name,
                "docstatus": 1
            },
            fields=["parent", "allocated_amount", "name"]
        )

        # Get Payment Entry details for each reference
        payments = []
        for ref in payment_references:
            pe = frappe.get_all(
                "Payment Entry",
                filters={"name": ref.parent, "docstatus": 1},
                fields=["name", "payment_date", "mode_of_payment", "posting_date"]
            )
            if pe:
                payments.append({
                    "mode_of_payment": pe[0].mode_of_payment,
                    "amount": ref.allocated_amount,
                    "payment_date": pe[0].payment_date or pe[0].posting_date,
                })

        # If there are payments, create a row for each payment
        if payments:
            for payment in payments:
                row_data = {
                    "customer": inv.customer,
                    "customer_name": inv.customer_name,
                    "custom_doc_no": inv.custom_doc_no or "",
                    "posting_date": inv.posting_date,
                    "due_date": inv.due_date,
                    "paid_amount": inv.grand_total or 0,
                    "outstanding_amount": inv.outstanding_amount or 0,
                    "mode_of_payment": payment.mode_of_payment or "",
                    "payment_amount": payment.amount or "",
                    "payment_date": payment.payment_date,
                }
                data.append(row_data)
        else:
            # If no payments, still show the invoice with empty payment info
            row_data = {
                "customer": inv.customer,
                "customer_name": inv.customer_name,
                "custom_doc_no": inv.custom_doc_no or "",
                "posting_date": inv.posting_date,
                "due_date": inv.due_date,
                "paid_amount": inv.grand_total or 0,
                "outstanding_amount": inv.outstanding_amount or 0,
                "mode_of_payment": "",
                "payment_amount": "",
                "payment_date": "",
            }
            data.append(row_data)

    return data


@frappe.whitelist()
def export_to_excel(filters):
    """
    Export laporan hasil tagihan to Excel with custom format
    One sheet per Sales Person
    """
    import io
    from frappe.utils.file_manager import save_file

    try:
        # Parse filters if string
        if isinstance(filters, str):
            filters = json.loads(filters)

        # Ensure filters is a dict
        if not isinstance(filters, dict):
            filters = {}

        # Parse sales_persons into a list
        sales_persons = filters.get("sales_persons")
        if sales_persons:
            if isinstance(sales_persons, str):
                try:
                    sales_persons = json.loads(sales_persons)
                except:
                    sales_persons = [sp.strip() for sp in sales_persons.split(",") if sp.strip()]
            if not isinstance(sales_persons, list):
                sales_persons = [sales_persons]
            sales_persons = [sp.strip() for sp in sales_persons if sp.strip()]
        else:
            sales_persons = []

        branch = filters.get("branch", "")
        sdate = filters.get("sdate", "")
        sdate_display = ""
        if sdate:
            try:
                sdate_display = getdate(sdate).strftime("%d-%m-%Y")
            except:
                sdate_display = str(sdate)

        # Define styles
        title_font = Font(bold=True, size=14)
        header_font_bold = Font(bold=True, size=10)
        header_font_normal = Font(size=9)
        small_font = Font(size=8)

        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_alignment = Alignment(horizontal="left", vertical="center")
        right_alignment = Alignment(horizontal="right", vertical="center")

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        thick_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thick"),
            bottom=Side(style="thick")
        )

        def build_sheet(ws, sp_name, data):
            """Build a single sheet for one Sales Person"""
            # Set page setup
            ws.page_setup.orientation = "landscape"
            ws.page_setup.paper_size = ws.PAPERSIZE_A4
            ws.page_setup.fitToPage = False

            # Set column widths
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 35
            ws.column_dimensions['D'].width = 8
            ws.column_dimensions['E'].width = 14
            ws.column_dimensions['F'].width = 11
            ws.column_dimensions['G'].width = 11
            ws.column_dimensions['H'].width = 13
            ws.column_dimensions['I'].width = 13
            ws.column_dimensions['J'].width = 8
            ws.column_dimensions['K'].width = 12
            ws.column_dimensions['L'].width = 10
            ws.column_dimensions['M'].width = 12
            ws.column_dimensions['N'].width = 12
            ws.column_dimensions['O'].width = 20

            # ===== ROW 2: Company Name and Sales Person =====
            branch_part = f" {branch}" if branch else ""
            ws['A2'] = f"PT. MHG{branch_part.upper()}"
            ws['A2'].font = header_font_bold
            ws['A2'].alignment = left_alignment

            # Sales Person label - M2
            ws['M2'] = "Sales Person"
            ws['M2'].font = small_font
            ws['N2'] = ":" + sp_name
            ws['N2'].font = small_font
            # ws['O2'] = sp_name
            # ws['O2'].font = small_font
            # ws['O2'].alignment = left_alignment

            # ===== ROW 3: No and Title =====
            ws['A3'] = "No:"
            ws['A3'].font = small_font
            ws['A3'].alignment = left_alignment

            # B3: date format mmyyyy.dd.A
            if sdate:
                try:
                    sdate_parsed = getdate(sdate)
                    ws['B3'] = sdate_parsed.strftime("%m%Y.%d.A")
                except:
                    ws['B3'] = ""
            else:
                ws['B3'] = ""
            ws['B3'].font = small_font
            ws['B3'].alignment = left_alignment

            ws.merge_cells('F3:I3')
            ws['F3'] = "LAPORAN HASIL TAGIHAN"
            ws['F3'].font = title_font
            ws['F3'].alignment = center_alignment

            # Tanggal - M3
            ws['M3'] = "Tanggal"
            ws['M3'].font = small_font
            ws['N3'] = ": " + sdate_display if sdate_display else ""
            ws['N3'].font = small_font
            ws['N3'].alignment = left_alignment

            # ===== ROW 5: Main Section Headers =====
            ws.merge_cells('A5:D5')
            ws['A5'] = "CUSTOMER"
            ws['A5'].font = header_font_bold
            ws['A5'].alignment = center_alignment
            for col in 'ABCD':
                ws[f'{col}5'].border = thin_border

            ws.merge_cells('E5:I5')
            ws['E5'] = "INFO TAGIHAN"
            ws['E5'].font = header_font_bold
            ws['E5'].alignment = center_alignment
            for col in 'EFGHI':
                ws[f'{col}5'].border = thin_border

            ws.merge_cells('J5:N5')
            ws['J5'] = "PEMBAYARAN"
            ws['J5'].font = header_font_bold
            ws['J5'].alignment = center_alignment
            for col in 'JKLMN':
                ws[f'{col}5'].border = thin_border

            ws.merge_cells('O5:O7')
            ws['O5'] = "Keterangan"
            ws['O5'].font = header_font_bold
            ws['O5'].alignment = center_alignment
            ws['O5'].border = thin_border
            ws['O6'].border = thin_border
            ws['O7'].border = thin_border

            # ===== ROW 6: Sub-headers =====
            ws['A6'] = "No."
            ws['A6'].font = header_font_bold
            ws['A6'].alignment = center_alignment
            ws['A6'].border = thin_border

            ws['B6'] = "Cust ID"
            ws['B6'].font = header_font_bold
            ws['B6'].alignment = center_alignment
            ws['B6'].border = thin_border

            ws['C6'] = "Nama Customer"
            ws['C6'].font = header_font_bold
            ws['C6'].alignment = center_alignment
            ws['C6'].border = thin_border

            ws['D6'] = "KET"
            ws['D6'].font = header_font_bold
            ws['D6'].alignment = center_alignment
            ws['D6'].border = thin_border

            ws.merge_cells('E6:E7')
            ws['E6'] = "Nomor"
            ws['E6'].font = header_font_bold
            ws['E6'].alignment = center_alignment
            ws['E6'].border = thin_border
            ws['E7'].border = thin_border

            ws.merge_cells('F6:G6')
            ws['F6'] = "Tanggal"
            ws['F6'].font = header_font_bold
            ws['F6'].alignment = center_alignment
            ws['F6'].border = thin_border

            ws.merge_cells('H6:H7')
            ws['H6'] = "Grand\nTotal"
            ws['H6'].font = header_font_bold
            ws['H6'].alignment = center_alignment
            ws['H6'].border = thin_border
            ws['H7'].border = thin_border

            ws.merge_cells('I6:I7')
            ws['I6'] = "Balance\nDue"
            ws['I6'].font = header_font_bold
            ws['I6'].alignment = center_alignment
            ws['I6'].border = thin_border
            ws['I7'].border = thin_border

            ws.merge_cells('J6:J7')
            ws['J6'] = "Bank"
            ws['J6'].font = header_font_bold
            ws['J6'].alignment = center_alignment
            ws['J6'].border = thin_border
            ws['J7'].border = thin_border

            ws.merge_cells('K6:M6')
            ws['K6'] = "Cek / BG Number"
            ws['K6'].font = header_font_bold
            ws['K6'].alignment = center_alignment
            ws['K6'].border = thin_border

            ws.merge_cells('N6:N7')
            ws['N6'] = "TUNAI"
            ws['N6'].font = header_font_bold
            ws['N6'].alignment = center_alignment
            ws['N6'].border = thin_border
            ws['N7'].border = thin_border

            # ===== ROW 7: Sub-sub-headers =====
            ws['F7'] = "Invoice"
            ws['F7'].font = header_font_normal
            ws['F7'].alignment = center_alignment
            ws['F7'].border = thin_border

            ws['G7'] = "J.T."
            ws['G7'].font = header_font_normal
            ws['G7'].alignment = center_alignment
            ws['G7'].border = thin_border

            ws['K7'] = "Nomor"
            ws['K7'].font = header_font_normal
            ws['K7'].alignment = center_alignment
            ws['K7'].border = thin_border

            ws['L7'] = "Tg.Jl"
            ws['L7'].font = header_font_normal
            ws['L7'].alignment = center_alignment
            ws['L7'].border = thin_border

            ws['M7'] = "Nominal"
            ws['M7'].font = header_font_normal
            ws['M7'].alignment = center_alignment
            ws['M7'].border = thin_border

            ws['E7'].border = thin_border
            ws['H7'].border = thin_border
            ws['I7'].border = thin_border
            ws['J7'].border = thin_border
            ws['N7'].border = thin_border

            # ===== DATA ROWS (starting row 8) =====
            start_row = 8
            for row_idx, row in enumerate(data, start=0):
                row_num = start_row + row_idx

                def set_cell(col, value, alignment=center_alignment, number_format=None, font=None):
                    cell = ws.cell(row=row_num, column=col, value=value)
                    cell.border = thin_border
                    cell.alignment = alignment
                    if number_format:
                        cell.number_format = number_format
                    if font:
                        cell.font = font

                set_cell(1, row_idx + 1, center_alignment)
                set_cell(2, row.get("customer", ""), left_alignment)
                set_cell(3, row.get("customer_name", ""), left_alignment)
                set_cell(4, "", center_alignment)
                set_cell(5, row.get("custom_doc_no", ""), center_alignment)

                inv_date = row.get("posting_date", "")
                if inv_date:
                    try:
                        inv_date = getdate(inv_date).strftime("%d/%m/%y")
                    except:
                        pass
                set_cell(6, inv_date, center_alignment)

                due_date = row.get("due_date", "")
                if due_date:
                    try:
                        due_date = getdate(due_date).strftime("%d/%m/%y")
                    except:
                        pass
                set_cell(7, due_date, center_alignment)

                grand_total = flt(row.get("paid_amount", 0))
                set_cell(8, grand_total, right_alignment, '#,##0')

                balance = flt(row.get("outstanding_amount", 0))
                set_cell(9, balance, right_alignment, '#,##0')

                set_cell(10, "", center_alignment)
                set_cell(11, "", left_alignment)
                set_cell(12, "", center_alignment)
                set_cell(13, "", right_alignment, '#,##0')
                set_cell(14, "", right_alignment, '#,##0')
                set_cell(15, "", left_alignment)

            # ===== TOTAL ROW =====
            total_row = start_row + len(data)

            ws.merge_cells(f'E{total_row}:G{total_row}')
            total_cell = ws.cell(row=total_row, column=5, value="TOTAL")
            total_cell.font = header_font_bold
            total_cell.alignment = right_alignment
            total_cell.border = thin_border

            for col in range(1, 16):
                ws.cell(row=total_row, column=col).border = thick_border

            total_grand = sum(flt(row.get("paid_amount", 0)) for row in data)
            ws.cell(row=total_row, column=8, value=total_grand).font = header_font_bold
            ws.cell(row=total_row, column=8).number_format = '#,##0'
            ws.cell(row=total_row, column=8).alignment = right_alignment

            total_balance = sum(flt(row.get("outstanding_amount", 0)) for row in data)
            ws.cell(row=total_row, column=9, value=total_balance).font = header_font_bold
            ws.cell(row=total_row, column=9).number_format = '#,##0'
            ws.cell(row=total_row, column=9).alignment = right_alignment

            ws.cell(row=total_row, column=14, value=0).font = header_font_bold
            ws.cell(row=total_row, column=14).number_format = '#,##0'
            ws.cell(row=total_row, column=14).alignment = right_alignment

            # ===== SIGNATURE SECTION =====
            sign_row = total_row + 2

            signatures = [
                ('B', 'Disiapkan Oleh,', 'Admin A/R'),
                ('D', 'Dicek Oleh,', 'Collector'),
                ('F', 'Mengetahui,', 'Koord. Finance'),
                ('H', 'Diterbitkan Oleh,', sp_name),
                ('J', 'Diterima Oleh,', 'Admin A/R'),
                ('L', 'Menyetujui,', 'BM/Assistant')
            ]

            for col, label, role in signatures:
                ws[f'{col}{sign_row}'] = label
                ws[f'{col}{sign_row}'].font = small_font
                ws[f'{col}{sign_row}'].alignment = left_alignment

                ws[f'{col}{sign_row+2}'] = role
                ws[f'{col}{sign_row+2}'].font = small_font
                ws[f'{col}{sign_row+2}'].alignment = left_alignment

        # Create Workbook with one sheet per Sales Person
        wb = Workbook()

        if sales_persons:
            for idx, sp in enumerate(sales_persons):
                # Get data for this specific sales person
                sp_filters = {**filters, "sales_persons": [sp]}
                sp_data = get_data(sp_filters)

                # Sanitize sheet name (max 31 chars, no special chars)
                sheet_name = sp[:31].replace("/", "-").replace("\\", "-").replace(":", "-").replace("?", "-").replace("*", "-")
                if idx == 0:
                    ws = wb.active
                    ws.title = sheet_name
                else:
                    ws = wb.create_sheet(title=sheet_name)

                build_sheet(ws, sp, sp_data)
        else:
            # Fallback: single sheet with all data
            ws = wb.active
            ws.title = "Laporan Hasil Tagihan"
            all_data = get_data(filters)
            sp_label = ", ".join(sales_persons) if sales_persons else ""
            build_sheet(ws, sp_label, all_data)

        # Save to file
        temp_file = io.BytesIO()
        wb.save(temp_file)
        temp_file.seek(0)

        file_name = f"Laporan_Tagihan_{now_datetime().strftime('%Y%m%d_%H%M%S')}.xlsx"

        # Save using Frappe file manager
        file_doc = save_file(
            fname=file_name,
            content=temp_file.getvalue(),
            dt="Report",
            dn="Laporan Hasil Tagihan",
            folder="Home/Attachments",
            decode=False
        )

        return {
            "file_url": file_doc.file_url,
            "file_name": file_name
        }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Laporan Hasil Tagihan Export Error")
        frappe.throw(f"Export failed: {str(e)}")
