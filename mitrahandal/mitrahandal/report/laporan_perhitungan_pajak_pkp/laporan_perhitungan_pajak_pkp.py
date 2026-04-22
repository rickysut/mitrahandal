# Copyright (c) 2026, Mitrahandal and contributors
# For license information, please see license.txt

import frappe
from frappe import _
from collections import defaultdict
from frappe.utils import getdate, format_date, fmt_money, nowdate, flt, now_datetime
import os
from datetime import datetime
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter



# Setup logging to mitrahandal/logs folder
log_dir = os.path.join(os.path.dirname(__file__), "..", "..", "..", "logs")
os.makedirs(log_dir, exist_ok=True)
log_file = os.path.join(log_dir, "laporan_perhitungan_pajak-" + now_datetime().strftime("%Y-%m-%d") + ".log")


def log_debug(message):
    timestamp = now_datetime().strftime("%Y-%m-%d %H:%M:%S")
    with open(log_file, "a") as f:
        f.write(f"[{timestamp}] {message}\n")

def execute(filters=None):
    columns = get_columns()
    data = get_data(filters)

    return columns, data

def validate_filters(filters):
    # Validate from_date and to_date
    from_date = filters.get("from_date")
    to_date = filters.get("to_date")

    if from_date and to_date:
        if from_date > to_date:
            frappe.throw(_("From Date cannot be greater than To Date"))

def get_columns():
    return [
        {
            "label": "TGL. INV",
            "fieldname": "inv_date",
            "fieldtype": "Data",
            "width": 100
        },
        {
            "label": "NO INV",
            "fieldname": "inv_number",
            "fieldtype": "Data",
            "width": 100
        },
        {
            "label": "NO CUST",
            "fieldname": "customer",
            "fieldtype": "Link",
            "options": "Customer",
            "width": 120
        },
        {
            "label": "PELANGGAN",
            "fieldname": "customer_name",
            "fieldtype": "Data",
            "width": 220
        },
        {
            "label": "NPWP",
            "fieldname": "customer_tax",
            "fieldtype": "Data",
            "width": 120
        },
        {
            "label": "NO. FAKTUR PAJAK",
            "fieldname": "coretax_number",
            "fieldtype": "Data",
            "width": 120
        },
        {
            "label": "TGL.FPJK",
            "fieldname": "coretax_date",
            "fieldtype": "Date",
            "width": 140
        },
        {
            "label": "DPP",
            "fieldname": "inv_dpp",
            "fieldtype": "Currency",
            "width": 140
        },
        {
            "label": "PPN",
            "fieldname": "inv_ppn",
            "fieldtype": "Currency",
            "width": 140
        },
    ]

def get_data(filters):
    log_debug(f"Filters received: {filters} Laporan Perhitungan Pajak Debug")
    validate_filters(filters)

    # Build filters for frappe.get_all
    si_filters = {"docstatus": 1}
    if filters.get("from_date") and filters.get("to_date"):
        si_filters["posting_date"] = [
            "between",
            [filters.get("from_date"), filters.get("to_date")]
        ]
        log_debug(f"Date filter applied: {filters.get('from_date')} to {filters.get('to_date')}")

    # Filter Company (required)
    if filters.get("company"):
        si_filters["company"] = filters.get("company")

    # Filter Customer (optional)
    if filters.get("customer"):
        si_filters["customer"] = filters.get("customer")

    log_debug(f"Filters: {si_filters}")

    # Get Sales Invoice list
    invoice_list = frappe.get_all(
        "Sales Invoice",
        filters={**si_filters, "status": "Paid"},
        fields=[
            "name",
            "customer",
            "customer_name",
            "custom_doc_no",
            "posting_date",
            "custom_doc_no",
            "tax_id",
            "net_total",
            "total_taxes_and_charges"
        ],
        order_by="posting_date ASC, name ASC",
    )
    log_debug(f"invoice list {invoice_list}")

    data = []
    for inv in invoice_list:
        # Get Item child table for warehouse filter
        items = frappe.get_all(
            "Sales Invoice Item",
            filters={"parent": inv.name},
            fields=["warehouse"],
        )

        # Apply warehouse filter if specified
        if filters.get("warehouse"):
            warehouse_match = any(item.warehouse == filters.get("warehouse") for item in items)
            if not warehouse_match:
                continue

                row_data = {
                    "inv_date": inv.posting_date,
                    "inv_number": inv.custom_doc_no,
                    "customer": inv.customer,
                    "customer_name": inv.customer_name,
                    "customer_tax": inv.tax_id,
                    "coretax_number": "0.001-" + inv.custom_doc_no,
                    "coretax_date": inv.posting_date,
                    "inv_dpp": inv.net_total,
                    "inv_ppn": inv.total_taxes_and_charges,
                }
                data.append(row_data)

    return data

@frappe.whitelist()
def export_to_excel(filters):
    """
    Export laporan perhitungan pajak pkp to Excel with custom format
    Format sesuai template screenshot
    """
    import io
    from frappe.utils.file_manager import save_file

    try:
        # Parse filters if string
        if isinstance(filters, str):
            filters = json.loads(filters)

        # Ensure filters is a dict
        if not isinstance(filters, dict):
            filters = {}

        # Get data
        data = get_data(filters)

        # Create Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "PT"

        # Define styles
        title_font = Font(bold=True, size=12)
        header_font_bold = Font(bold=True, size=10)
        small_font = Font(size=9)
        small_font_bold = Font(bold=True, size=9)

        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        left_alignment = Alignment(horizontal="left", vertical="center")
        right_alignment = Alignment(horizontal="right", vertical="center")

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        thick_border_bottom = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thick")
        )

        # Set page setup for print
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paper_size = ws.PAPERSIZE_A4
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToPagesWide = 1
        ws.page_setup.fitToPagesTall = 99

        # Set column widths (based on screenshot)
        ws.column_dimensions['A'].width = 5      # NO.
        ws.column_dimensions['B'].width = 12     # TGL. INV
        ws.column_dimensions['C'].width = 12     # NO INV
        ws.column_dimensions['D'].width = 12     # NO CUST
        ws.column_dimensions['E'].width = 35     # PELANGGAN
        ws.column_dimensions['F'].width = 18     # NPWP
        ws.column_dimensions['G'].width = 14     # NO. FAKTUR PAJAK
        ws.column_dimensions['H'].width = 12     # TGL. FPJK
        ws.column_dimensions['I'].width = 14     # DPP
        ws.column_dimensions['J'].width = 14     # PPN

        # ===== ROW 1: Header =====
        # Top left: timestamp (nowdate)
        current_datetime = now_datetime().strftime("%d/%m/%Y %H:%M")
        ws['A1'] = current_datetime
        ws['A1'].font = small_font
        ws['A1'].alignment = left_alignment

        # Top right: Halaman
        total_pages = max(1, (len(data) // 50) + 1)
        ws['J1'] = f"Halaman 1 dari {total_pages}"
        ws['J1'].font = small_font
        ws['J1'].alignment = right_alignment

        # ===== ROW 2: Company Name =====
        company_name = "PT. MITRA HANDAL GEMILANG JAKARTA"
        ws.merge_cells('A2:J2')
        ws['A2'] = company_name
        ws['A2'].font = title_font
        ws['A2'].alignment = center_alignment

        # ===== ROW 3: Report Title =====
        # Get month and year from filters
        from_date = filters.get("from_date")
        if from_date:
            month_year = getdate(from_date).strftime("%B %Y").upper()
        else:
            month_year = now_datetime().strftime("%B %Y").upper()
        
        ws.merge_cells('A3:J3')
        ws['A3'] = f"LAPORAN PERHITUNGAN PAJAK PKP MASA PAJAK BULAN {month_year}"
        ws['A3'].font = header_font_bold
        ws['A3'].alignment = center_alignment

        # ===== ROW 4: Empty =====
        # Skip row 4 for spacing

        # ===== ROW 5: Column Headers =====
        headers = [
            ("NO.", "A5"),
            ("TGL. INV", "B5"),
            ("NO INV", "C5"),
            ("NO CUST", "D5"),
            ("PELANGGAN", "E5"),
            ("NPWP", "F5"),
            ("NO. FAKTUR PAJAK", "G5"),
            ("TGL. FPJK", "H5"),
            ("DPP", "I5"),
            ("PPN", "J5"),
        ]
        
        for label, cell in headers:
            ws[cell] = label
            ws[cell].font = header_font_bold
            ws[cell].alignment = center_alignment
            ws[cell].border = thin_border

        # ===== DATA ROWS (starting row 6) =====
        start_row = 6
        for row_idx, row in enumerate(data, start=0):
            row_num = start_row + row_idx

            # Helper function to set cell with border and alignment
            def set_cell(col_letter, value, alignment=center_alignment, number_format=None, font=None):
                cell = ws[f'{col_letter}{row_num}']
                cell.value = value
                cell.border = thin_border
                cell.alignment = alignment
                if number_format:
                    cell.number_format = number_format
                if font:
                    cell.font = font

            # NO.
            set_cell('A', row_idx + 1, center_alignment)

            # TGL. INV
            inv_date = row.get("inv_date", "")
            if inv_date:
                try:
                    inv_date = getdate(inv_date).strftime("%d/%m/%Y")
                except:
                    pass
            set_cell('B', inv_date, center_alignment)

            # NO INV
            set_cell('C', row.get("inv_number", ""), center_alignment)

            # NO CUST
            set_cell('D', row.get("customer", ""), left_alignment)

            # PELANGGAN
            set_cell('E', row.get("customer_name", ""), left_alignment)

            # NPWP
            set_cell('F', row.get("customer_tax", ""), left_alignment)

            # NO. FAKTUR PAJAK
            set_cell('G', row.get("coretax_number", ""), left_alignment)

            # TGL. FPJK
            coretax_date = row.get("coretax_date", "")
            if coretax_date:
                try:
                    coretax_date = getdate(coretax_date).strftime("%d/%m/%Y")
                except:
                    pass
            set_cell('H', coretax_date, center_alignment)

            # DPP
            dpp = flt(row.get("inv_dpp", 0))
            set_cell('I', dpp, right_alignment, '#,##0')

            # PPN
            ppn = flt(row.get("inv_ppn", 0))
            set_cell('J', ppn, right_alignment, '#,##0')

        # ===== TOTAL ROW =====
        total_row = start_row + len(data)

        # Merge A:G for TOTAL label
        ws.merge_cells(f'A{total_row}:G{total_row}')
        total_cell = ws[f'A{total_row}']
        total_cell.value = "T O T A L"
        total_cell.font = header_font_bold
        total_cell.alignment = center_alignment
        total_cell.border = thick_border_bottom

        # Add thick border bottom to all cells in total row
        for col_letter in 'ABCDEFGHIJ':
            cell = ws[f'{col_letter}{total_row}']
            cell.border = thick_border_bottom

        # Total DPP
        total_dpp = sum(flt(row.get("inv_dpp", 0)) for row in data)
        total_dpp_cell = ws[f'I{total_row}']
        total_dpp_cell.value = total_dpp
        total_dpp_cell.font = header_font_bold
        total_dpp_cell.number_format = '#,##0'
        total_dpp_cell.alignment = right_alignment

        # Total PPN
        total_ppn = sum(flt(row.get("inv_ppn", 0)) for row in data)
        total_ppn_cell = ws[f'J{total_row}']
        total_ppn_cell.value = total_ppn
        total_ppn_cell.font = header_font_bold
        total_ppn_cell.number_format = '#,##0'
        total_ppn_cell.alignment = right_alignment

        # Save to file
        temp_file = io.BytesIO()
        wb.save(temp_file)
        temp_file.seek(0)

        file_name = f"Laporan_Pajak_PKP_{now_datetime().strftime('%Y%m%d_%H%M%S')}.xlsx"

        # Save using Frappe file manager
        file_doc = save_file(
            fname=file_name,
            content=temp_file.getvalue(),
            dt="Report",
            dn="Laporan Perhitungan Pajak PKP",
            folder="Home/Attachments",
            decode=False
        )

        return {
            "file_url": file_doc.file_url,
            "file_name": file_name
        }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Laporan Perhitungan Pajak PKP Export Error")
        frappe.throw(f"Export failed: {str(e)}")